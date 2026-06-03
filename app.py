import io, os, sys, re
import streamlit as st
import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

st.set_page_config(page_title="Proyecto Juliana", layout="wide", initial_sidebar_state="collapsed")

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');
    html, body, [class*="css"]  { font-family: 'Inter', sans-serif; }
    .stApp { background: #f0f2f6; }
    .main > div { padding: 1rem 2rem; }
    .header { background: linear-gradient(135deg, #1e3a5f 0%, #2d5a8e 100%); padding: 1.5rem 2rem; border-radius: 16px; margin: -1rem -2rem 1.5rem -2rem; color: white; }
    .header h1 { margin: 0; font-size: 1.8rem; font-weight: 800; letter-spacing: -0.5px; text-align: left; }
    .header p { margin: 0.3rem 0 0 0; opacity: 0.8; font-size: 0.9rem; text-align: left; }
    .card { background: white; border-radius: 12px; padding: 1.2rem 1.5rem; box-shadow: 0 1px 3px rgba(0,0,0,0.08); margin-bottom: 1rem; }
    .card-title { font-size: 0.75rem; font-weight: 600; text-transform: uppercase; letter-spacing: 0.5px; color: #888; margin-bottom: 0.5rem; }
    .stat-value { font-size: 1.6rem; font-weight: 700; color: #1e3a5f; }
    .stat-label { font-size: 0.75rem; color: #888; font-weight: 500; }
    .cole-header { font-size: 1.1rem; font-weight: 700; color: #1e3a5f; margin-bottom: 0.3rem; }
    .cole-dane { font-size: 0.8rem; color: #999; }
    .grade-badge { display: inline-block; background: #e8f0fe; color: #1e3a5f; font-weight: 600; font-size: 0.85rem; padding: 0.2rem 0.7rem; border-radius: 20px; }
    .badge-complete { background: #d4edda; color: #155724; font-weight: 600; font-size: 0.75rem; padding: 0.15rem 0.5rem; border-radius: 12px; }
    .badge-incomplete { background: #f8d7da; color: #721c24; font-weight: 600; font-size: 0.75rem; padding: 0.15rem 0.5rem; border-radius: 12px; }
    .element-container:has(.stRadio) { display: flex !important; justify-content: center !important; width: 100% !important; }
    .row-widget.stRadio { display: flex !important; justify-content: center !important; width: 100% !important; margin: 0.8rem 0 1.5rem 0 !important; }
    .stRadio { display: flex !important; justify-content: center !important; width: 100% !important; }
    .stRadio > div { display: flex !important; justify-content: center !important; width: 100% !important; }
    .stRadio [role="radiogroup"] { background: linear-gradient(135deg, #e8f0fe 0%, #ffffff 100%) !important; border-radius: 12px !important; padding: 6px !important; display: flex !important; gap: 6px !important; width: auto !important; border: none !important; box-shadow: 0 2px 6px rgba(30,58,95,0.08) !important; border-left: 4px solid #1e3a5f !important; }
    .stRadio [role="radiogroup"] label { flex: 0 0 auto !important; text-align: center !important; padding: 0.5rem 2rem !important; border-radius: 9px !important; font-weight: 700 !important; font-size: 0.95rem !important; color: #4a6a8a !important; margin: 0 !important; transition: all 0.15s !important; }
    .stRadio [role="radiogroup"] label:hover { background: rgba(30,58,95,0.06) !important; }
    .stRadio [role="radiogroup"] input { accent-color: #1e3a5f !important; }
    .stRadio [role="radiogroup"] input:checked ~ div { background: #1e3a5f !important; color: white !important; border-radius: 9px !important; box-shadow: 0 2px 4px rgba(30,58,95,0.15) !important; }
    div[data-testid="stMetric"] { background: white; border-radius: 10px; padding: 0.8rem; box-shadow: 0 1px 3px rgba(0,0,0,0.05); }
    div[data-testid="stMetric"] label { font-size: 0.75rem; font-weight: 500; color: #888; }
    div[data-testid="stMetric"] div[data-testid="stMetricValue"] { font-size: 1.5rem; font-weight: 700; color: #1e3a5f; }
    .stDataFrame { border-radius: 10px; overflow: hidden; box-shadow: 0 1px 3px rgba(0,0,0,0.05); }
    .stDataFrame td, .stDataFrame th { font-size: 0.75rem; }
    .inc-alert { background: #fff5f5; border-left: 4px solid #e53e3e; border-radius: 8px; padding: 0.8rem 1rem; margin: 0.5rem 0; }
    .stExpander { background: white; border-radius: 10px; box-shadow: 0 1px 3px rgba(0,0,0,0.05); border: none !important; margin-bottom: 0.6rem; }
    .stExpander summary { font-weight: 600; padding: 0.6rem 1rem; }
    .stExpander summary span { font-size: 0.9rem; }
    section[data-testid="stExpander"] { border: none !important; }
    .st-ae { border: none !important; }
    .stAlert { border-radius: 8px; }
    hr { margin: 1.5rem 0; opacity: 0.15; }
    .block-container { max-width: 1400px; padding: 1rem 2rem; margin: 0 auto; }
    h1, h2, h3, h4, h5, h6 { text-align: center; }
    .stMultiSelect { margin-bottom: 1rem; }
    button[kind="secondary"] { border-radius: 8px; font-weight: 500; }
    .stDownloadButton button { border-radius: 8px; background: #1e3a5f; color: white; font-weight: 500; }
    .stDownloadButton button:hover { background: #2d5a8e; }
    div[data-testid="stMetric"] { background: linear-gradient(135deg, #e8f0fe 0%, #ffffff 100%); border-radius: 10px; padding: 0.4rem 0.6rem; box-shadow: 0 2px 6px rgba(30,58,95,0.08); border-left: 3px solid #1e3a5f; height: 100%; display: flex; flex-direction: column; justify-content: center; }
    div[data-testid="stMetric"] label { font-size: 0.65rem; font-weight: 600; color: #4a6a8a; letter-spacing: 0.1px; white-space: normal !important; overflow: visible !important; margin-bottom: 0.1rem; }
    div[data-testid="stMetric"] div[data-testid="stMetricValue"] { font-size: 1.1rem; font-weight: 700; color: #1e3a5f; line-height: 1.1; }
    .stMultiSelect { margin-bottom: 1rem; }
    .stMultiSelect label { font-size: 1rem !important; font-weight: 700 !important; color: #1e3a5f !important; padding-top: 0.3rem !important; }
    .stMultiSelect [data-baseweb="select"] { min-height: 2.8rem !important; border-radius: 8px !important; }
    .stMultiSelect [data-baseweb="select"] span { background: #e8f0fe !important; color: #1e3a5f !important; font-size: 0.95rem !important; font-weight: 500 !important; border-radius: 4px !important; padding: 0.2rem 0.5rem !important; }
    .stMultiSelect [data-baseweb="select"] div { border-color: #c0d0e0 !important; border-radius: 8px !important; font-size: 0.95rem !important; }
    .stMultiSelect [data-baseweb="select"] input { font-size: 0.95rem !important; }
    .stMultiSelect [role="listbox"] li { font-size: 0.95rem !important; padding: 0.4rem 0.6rem !important; }
    .stMultiSelect [data-baseweb="tag"] { font-size: 0.85rem !important; padding: 0.2rem !important; }
    .element-container:has(.stMultiSelect) { display: flex; justify-content: center; }
    div[data-testid="column"]:has(.stMultiSelect) { max-width: 350px; margin: 0 auto; }

</style>
""", unsafe_allow_html=True)

st.markdown('<div class="header"><h1>Proyecto Juliana</h1><p>Sistema de seguimiento académico</p></div>', unsafe_allow_html=True)

RUTA_ACUM = os.path.join(os.path.dirname(__file__), "INFORME CLIENTE", "ACUMULADO GENERAL.xlsx")

# ─── LECTURA DE DATOS ────────────────────────────────────────────
@st.cache_data
def leer_acumulado():
    if not os.path.exists(RUTA_ACUM):
        return None
    wb = openpyxl.load_workbook(RUTA_ACUM)
    dfs = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        headers = [str(c.value) if c.value else f"COL{i}" for i, c in enumerate(ws[1])]
        eval_cols = [i for i, h in enumerate(headers) if h.endswith("_EVAL")]
        meta_nombres = ["TIPO", "NOMBRE SEDE", "CÓDIGO DANE SEDE", "NOMBRES ESTUDIANTE",
                         "CÓD. EST.", "GRADO", "CURSO", "PRUEBA"]
        meta_idx = [(i, headers.index(c)) for i, c in enumerate(meta_nombres) if c in headers]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(c is None or str(c).strip() == "" for c in row):
                continue
            correctas = sum(1 for i in eval_cols if i < len(row) and str(row[i]).strip().upper() == "CORRECTO")
            total = len(eval_cols)
            meta = {}
            for mi, col_i in meta_idx:
                v = row[col_i] if col_i < len(row) and row[col_i] is not None else ""
                val = re.sub(r'\s+', ' ', str(v).strip().upper())
                meta[meta_nombres[mi]] = val
            meta["_KEY"] = meta["CÓD. EST."] if meta["CÓD. EST."] else meta["NOMBRES ESTUDIANTE"]
            meta["CORRECTAS"] = correctas
            meta["TOTAL"] = total
            meta["%"] = round(correctas / total * 100, 1) if total else 0
            rows.append(meta)
        if rows:
            dfs.append(pd.DataFrame(rows))
    wb.close()
    return pd.concat(dfs, ignore_index=True) if dfs else None

# ─── PROCESAMIENTO ───────────────────────────────────────────────
def procesar_datos(df):
    if df is None or df.empty:
        return None, None
    # Si el archivo no tiene TIPO (datos viejos), crear columna vacía
    if "TIPO" not in df.columns:
        df["TIPO"] = ""
    df["_NAMEKEY"] = (df["NOMBRE SEDE"] + "|" + df["CÓDIGO DANE SEDE"] + "|" +
                      df["NOMBRES ESTUDIANTE"] + "|" + df["GRADO"].astype(str) + "|" + df["CURSO"].astype(str))

    # ── Resumen agrupado por sede + grado + materia ──
    df["INCORRECTAS"] = df["TOTAL"] - df["CORRECTAS"]
    resumen = df.groupby(["CÓDIGO DANE SEDE", "NOMBRE SEDE", "TIPO", "GRADO", "PRUEBA"]).agg(
        estudiantes=("_NAMEKEY", "nunique"),
        correctas=("CORRECTAS", "sum"),
        incorrectas=("INCORRECTAS", "sum"),
    ).reset_index()
    resumen["promedio_pct"] = resumen.apply(
        lambda r: round(r["correctas"] / (r["correctas"] + r["incorrectas"]) * 100, 1)
        if (r["correctas"] + r["incorrectas"]) > 0 else 0, axis=1
    )
    resumen["prom_correctas"] = resumen.apply(
        lambda r: round(r["correctas"] / r["estudiantes"], 1) if r["estudiantes"] > 0 else 0, axis=1
    )
    resumen["prom_incorrectas"] = resumen.apply(
        lambda r: round(r["incorrectas"] / r["estudiantes"], 1) if r["estudiantes"] > 0 else 0, axis=1
    )
    resumen["_GRADO_NUM"] = pd.to_numeric(resumen["GRADO"], errors="coerce")
    resumen = resumen.sort_values(["CÓDIGO DANE SEDE", "_GRADO_NUM", "PRUEBA"]).reset_index(drop=True).drop(columns=["_GRADO_NUM"])

    return {
        "df": df,
        "resumen": resumen,
    }

# ─── INICIO ──────────────────────────────────────────────────────
df = leer_acumulado()
datos = procesar_datos(df)

if datos is None:
    st.info("No hay datos disponibles. Vuelve pronto.")
    st.stop()

DF = datos["df"]
RES = datos["resumen"]

# ─── NAVEGACIÓN ──────────────────────────────────────────────────
nivel = st.radio("Vista", ["Resumen", "Base General"],
                 horizontal=True, label_visibility="collapsed")

# ══════════════════════════════════════════════════════════════════
# RESUMEN (VISOR PRINCIPAL)
# ══════════════════════════════════════════════════════════════════
if nivel == "Resumen":
    st.markdown("### Resumen")

    # ── Filters ─────────────────────────────────────────────────
    opciones_tipo = sorted(RES["TIPO"].dropna().unique().tolist())
    opciones_grado = sorted(RES["GRADO"].astype(str).unique().tolist(), key=lambda g: int(g))

    if "filtro_tipo" not in st.session_state:
        st.session_state.filtro_tipo = opciones_tipo
    if "filtro_grado" not in st.session_state:
        st.session_state.filtro_grado = opciones_grado

    c1, c2 = st.columns(2)
    with c1:
        st.multiselect("TIPO", opciones_tipo, default=st.session_state.filtro_tipo,
                       key="filtro_tipo")
    with c2:
        st.multiselect("GRADO", opciones_grado, default=st.session_state.filtro_grado,
                       key="filtro_grado")

    filtros = (
        RES["TIPO"].isin(st.session_state.filtro_tipo) &
        RES["GRADO"].astype(str).isin(st.session_state.filtro_grado)
    )
    RES_f = RES[filtros].copy()

    total_filas = len(RES_f)
    total_colegios = RES_f["CÓDIGO DANE SEDE"].nunique()
    RES_f["_PRUEBA_NORM"] = RES_f["PRUEBA"].str.replace("Á", "A").str.replace("É", "E").str.replace("Í", "I").str.replace("Ó", "O").str.replace("Ú", "U")

    def total_est(df, materia_norm):
        sub = df[df["_PRUEBA_NORM"] == materia_norm]
        return int(sub["estudiantes"].sum())

    def calc_pct(df, materia_norm):
        sub = df[df["_PRUEBA_NORM"] == materia_norm]
        c = sub["correctas"].sum()
        i = sub["incorrectas"].sum()
        return round(c / (c + i) * 100, 1) if (c + i) > 0 else 0

    total_est_len = total_est(RES_f, "LENGUAJE")
    total_est_mat = total_est(RES_f, "MATEMATICAS")
    pct_len = calc_pct(RES_f, "LENGUAJE")
    pct_mat = calc_pct(RES_f, "MATEMATICAS")

    a, b, c, d, e, f = st.columns(6)
    a.metric("Filas", total_filas)
    b.metric("Colegios", total_colegios)
    c.metric("Est. Lenguaje", total_est_len)
    d.metric("Est. Matemáticas", total_est_mat)
    e.metric("% Lenguaje", f"{pct_len}%")
    f.metric("% Matemáticas", f"{pct_mat}%")

    # ── % Acierto por TIPO y Materia ──
    def pct_tipo_materia(df, tipo, materia_norm):
        sub = df[(df["TIPO"] == tipo) & (df["_PRUEBA_NORM"] == materia_norm)]
        c = sub["correctas"].sum()
        i = sub["incorrectas"].sum()
        return round(c / (c + i) * 100, 1) if (c + i) > 0 else 0
    tipos_disponibles = sorted(RES_f["TIPO"].dropna().unique().tolist())
    st.markdown("#### % Acierto por Tipo")
    cols = st.columns(len(tipos_disponibles) * 2 if tipos_disponibles else 2)
    idx = 0
    for t in tipos_disponibles:
        p_l = pct_tipo_materia(RES_f, t, "LENGUAJE")
        p_m = pct_tipo_materia(RES_f, t, "MATEMATICAS")
        cols[idx].metric(f"{t} - Lenguaje", f"{p_l}%")
        cols[idx+1].metric(f"{t} - Matemáticas", f"{p_m}%")
        idx += 2

    cols_mostrar = ["TIPO", "CÓDIGO DANE SEDE", "NOMBRE SEDE", "GRADO", "PRUEBA",
                     "estudiantes", "prom_correctas", "prom_incorrectas", "promedio_pct"]
    tabla = RES_f[cols_mostrar].copy()
    col_renombre = ["Tipo", "Código DANE", "Sede", "Grado", "Materia",
                     "Estudiantes", "Prom. Correctas", "Prom. Incorrectas", "% Desempeño"]
    tabla.columns = col_renombre
    st.dataframe(tabla, width="stretch", hide_index=True)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        tabla.to_excel(writer, index=False, sheet_name="Resumen")
    buf.seek(0)
    st.download_button("Descargar Resumen (.xlsx)", buf.read(),
                       "resumen_filtrado.xlsx")

# ══════════════════════════════════════════════════════════════════
# BASE GENERAL (DESCARGA)
# ══════════════════════════════════════════════════════════════════
elif nivel == "Base General":
    st.markdown("### Base General")

    cols_base = ["TIPO", "CÓDIGO DANE SEDE", "NOMBRE SEDE", "NOMBRES ESTUDIANTE",
                  "CÓD. EST.", "GRADO", "CURSO", "PRUEBA", "CORRECTAS", "TOTAL", "%"]
    mostrar = [c for c in cols_base if c in DF.columns]
    tabla = DF[mostrar].copy()
    st.dataframe(tabla, width="stretch", hide_index=True)

    if os.path.exists(RUTA_ACUM):
        wb = openpyxl.load_workbook(RUTA_ACUM)
        for ws in wb.worksheets:
            headers = [str(c.value) if c.value else f"COL{i}" for i, c in enumerate(ws[1])]
            eval_cols_idx = [i for i, h in enumerate(headers) if h.endswith("_EVAL")]
            # Encontrar columnas CORRECTAS, INCORRECTAS, PORCENTAJE ACIERTO
            col_corr = next((i for i, h in enumerate(headers) if h == "CORRECTAS"), None)
            col_inc = next((i for i, h in enumerate(headers) if h == "INCORRECTAS"), None)
            col_pct = next((i for i, h in enumerate(headers) if h == "PORCENTAJE ACIERTO"), None)
            if col_corr is None or col_inc is None or col_pct is None:
                continue
            total_eval = len(eval_cols_idx)
            if total_eval == 0:
                continue
            for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                correctas = sum(1 for i in eval_cols_idx if i < len(row) and str(row[i]).strip().upper() == "CORRECTO")
                incorrectas = total_eval - correctas
                ws.cell(row=r_idx, column=col_corr + 1).value = correctas
                ws.cell(row=r_idx, column=col_inc + 1).value = incorrectas
                ws.cell(row=r_idx, column=col_pct + 1).value = correctas / total_eval
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        wb.close()
        st.download_button(
            "⬇️ Descargar ACUMULADO GENERAL.xlsx",
            buf.read(),
            "ACUMULADO GENERAL.xlsx",
        )
