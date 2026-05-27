"""
CALIFICADOR LOCAL — solo califica sin subir a GitHub ni tocar acumulados.
Útil para probar/ver resultados sin afectar el sistema principal.
"""
import os, re, shutil
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

BASE = os.path.dirname(os.path.abspath(__file__))
RUTA_RESPUESTAS = os.path.join(BASE, "RESPUESTAS CORRECTAS PROYECTO JULIANA.xlsx")
DIR_ENTRADA = os.path.join(BASE, "Pruebas a calificar")
DIR_SALIDA = os.path.join(BASE, "Resultados locales")

MATERIA_NOMBRES = {"matematicas": "MATEMÁTICAS", "lenguaje": "LENGUAJE"}

COLUMNAS_ESTANDAR = [
    "TIPO", "CÓDIGO DANE SEDE", "NOMBRE SEDE", "NOMBRES ESTUDIANTE",
    "CÓD. EST.", "GRADO", "CURSO", "PRUEBA"
]

# ─── Estilos ──────────────────────────────────────────────────────
FILL_HEADER = PatternFill("solid", fgColor="1F3864")
FONT_HEADER = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
FILL_OK = PatternFill("solid", fgColor="C6EFCE")
FILL_FAIL = PatternFill("solid", fgColor="FFC7CE")
FONT_CELL = Font(size=10, name="Calibri")
FONT_SUM = Font(size=10, name="Calibri", bold=True)
ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

def sanitizar(s):
    return re.sub(r'[\\/*?:\[\]]', '', s).strip().upper()[:31]

def estilo_header(ws, ncols):
    for c in range(1, ncols + 1):
        celda = ws.cell(1, c)
        celda.fill = FILL_HEADER
        celda.font = FONT_HEADER
        celda.alignment = ALIGN
        celda.border = BORDER

def ancho_columnas(ws, cols):
    for i, c in enumerate(cols, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = max(len(c) + 3, 12)

PARED = re.compile(r'[^a-zA-Z0-9áéíóúÁÉÍÓÚñÑüÜ\s]')

def normalizar(v):
    return re.sub(r'\s+', ' ', PARED.sub('', str(v).strip().upper())).strip()

def detectar_columnas(ws):
    deteccion = {"NOMBRES ESTUDIANTE": None, "CÓDIGO DANE SEDE": None,
                 "NOMBRE SEDE": None, "CÓD. EST.": None,
                 "GRADO": None, "CURSO": None, "PRUEBA": None}
    header_map = {}
    for c in range(1, ws.max_column + 1):
        v = str(ws.cell(1, c).value or "").strip().upper()
        v = re.sub(r'\s+', ' ', v)
        header_map[c] = v
        for key in deteccion:
            if key in v:
                deteccion[key] = c
    ini_resp = None
    for c, h in header_map.items():
        if h.startswith("P") and h[1:2].isdigit():
            if not h.endswith("_EVAL"):
                ini_resp = c
                break
    if not ini_resp:
        for c in range(max(deteccion.values()) + 1 if any(deteccion.values()) else 1, ws.max_column + 1):
            ini_resp = c
            break
    return deteccion, ini_resp, header_map

SUJETO_MAP = {"lenguaje": "LENGUAJE", "lectura": "LENGUAJE", "matematicas": "MATEMÁTICAS"}

def cargar_respuestas():
    if not os.path.exists(RUTA_RESPUESTAS):
        return {}
    wb = openpyxl.load_workbook(RUTA_RESPUESTAS)
    respuestas = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            g = str(row[0]).strip()
            m = str(row[1]).strip().lower()
            for clave in SUJETO_MAP:
                if clave in m:
                    m = SUJETO_MAP[clave]
                    break
            p = str(row[2]).strip()
            r = str(row[3]).strip().upper() if row[3] else ""
            key = (g, m, p)
            respuestas[key] = r
    wb.close()
    return respuestas

def procesar(ruta_archivo, respuestas):
    wb = openpyxl.load_workbook(ruta_archivo)
    ws = wb.active
    cols, ini_resp, _ = detectar_columnas(ws)
    if cols["CÓDIGO DANE SEDE"] is None or cols["NOMBRES ESTUDIANTE"] is None:
        return []

    filas = list(ws.iter_rows(min_row=2, values_only=True))
    e_dane = cols["CÓDIGO DANE SEDE"]
    e_sede = cols["NOMBRE SEDE"]
    e_nombre = cols["NOMBRES ESTUDIANTE"]
    e_cod = cols["CÓD. EST."]
    e_grado = cols["GRADO"]
    e_curso = cols["CURSO"]
    e_prueba = cols["PRUEBA"]

    alumnos = []
    for fila in filas:
        if not fila[ini_resp - 1]:
            continue
        nombre = normalizar(fila[e_nombre - 1]) if e_nombre and len(fila) >= e_nombre else ""
        if not nombre:
            continue
        cod_dane = normalizar(fila[e_dane - 1]) if e_dane and len(fila) >= e_dane else ""
        sede = normalizar(fila[e_sede - 1]) if e_sede and len(fila) >= e_sede else ""
        cod = normalizar(fila[e_cod - 1]) if e_cod and len(fila) >= e_cod else ""
        grado = normalizar(fila[e_grado - 1]) if e_grado and len(fila) >= e_grado else ""
        curso = normalizar(fila[e_curso - 1]) if e_curso and len(fila) >= e_curso else ""
        materia_raw = normalizar(fila[e_prueba - 1]) if e_prueba and len(fila) >= e_prueba else ""
        materia = "LENGUAJE" if "LENGUAJE" in materia_raw or "LECTURA" in materia_raw else "MATEMÁTICAS"

        detalles = []
        for q in range(20):
            col = ini_resp + q - 1
            if col >= len(fila):
                break
            resp_alumno = normalizar(fila[col]) if fila[col] else ""
            clave = (grado, materia, f"P{q + 1:02d}")
            correcta = respuestas.get(clave, "")
            ok = False
            if resp_alumno:
                ok = resp_alumno == normalizar(correcta)
            detalles.append({"resp": resp_alumno or "", "ok": ok, "correcta": correcta})

        correctas = sum(1 for d in detalles if d["ok"])
        total = len(detalles)
        incorrectas = total - correctas
        pct = round(correctas / total * 100, 1) if total else 0

        alumnos.append({
            "cod_dane": cod_dane, "sede": sede, "nombre": nombre,
            "cod": cod, "grado": grado, "curso": curso, "materia": materia,
            "detalles": detalles, "correctas": correctas,
            "incorrectas": incorrectas, "total": total, "pct": pct
        })
    wb.close()
    return alumnos

def generar_excel(alumnos, ruta_salida):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultados"

    headers = COLUMNAS_ESTANDAR + [f"P{q:02d}" for q in range(1, 21)] + \
              [f"P{q:02d}_EVAL" for q in range(1, 21)] + \
              ["CORRECTAS", "INCORRECTAS", "TOTAL", "PORCENTAJE ACIERTO"]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    estilo_header(ws, len(headers))
    ancho_columnas(ws, headers)
    ws.freeze_panes = "A2"

    for i, al in enumerate(alumnos, 2):
        ws.cell(i, 1, al["cod_dane"]).font = FONT_CELL
        ws.cell(i, 2, al["sede"]).font = FONT_CELL
        ws.cell(i, 3, al["nombre"]).font = FONT_CELL
        ws.cell(i, 4, al["cod"]).font = FONT_CELL
        ws.cell(i, 5, al["grado"]).font = FONT_CELL
        ws.cell(i, 6, al["curso"]).font = FONT_CELL
        ws.cell(i, 7, al["materia"]).font = FONT_CELL
        ws.cell(i, 8, "").font = FONT_CELL
        for c in range(1, 9):
            ws.cell(i, c).alignment = ALIGN
            ws.cell(i, c).border = BORDER
        for q, det in enumerate(al["detalles"]):
            col_resp = 9 + q * 2
            col_eval = col_resp + 1
            ws.cell(i, col_resp, det["resp"]).font = FONT_CELL
            ws.cell(i, col_resp).alignment = ALIGN
            ws.cell(i, col_resp).border = BORDER
            celda = ws.cell(i, col_eval, "CORRECTO" if det["ok"] else "INCORRECTO")
            celda.font = FONT_CELL
            celda.alignment = ALIGN
            celda.border = BORDER
            celda.fill = FILL_OK if det["ok"] else FILL_FAIL
        ws.cell(i, 49, al["correctas"]).font = FONT_SUM
        ws.cell(i, 50, al["incorrectas"]).font = FONT_SUM
        ws.cell(i, 51, al["total"]).font = FONT_SUM
        ws.cell(i, 52, f"{al['pct']}%").font = FONT_SUM
        for c in range(49, 53):
            ws.cell(i, c).alignment = ALIGN
            ws.cell(i, c).border = BORDER

    wb.save(ruta_salida)
    wb.close()

def main():
    print("=" * 55)
    print("  CALIFICADOR LOCAL (sin subir a GitHub)")
    print("=" * 55)

    if not os.path.exists(RUTA_RESPUESTAS):
        print(f"  [!] No se encuentra: RESPUESTAS CORRECTAS PROYECTO JULIANA.xlsx")
        return

    respuestas = cargar_respuestas()
    print(f"  + Respuestas cargadas: {len(respuestas)} combinaciones")

    archivos = sorted(
        os.path.join(DIR_ENTRADA, f)
        for f in os.listdir(DIR_ENTRADA)
        if f.lower().endswith(".xlsx") and os.path.isfile(os.path.join(DIR_ENTRADA, f))
    )
    if not archivos:
        print("  [!] No hay archivos .xlsx en 'Pruebas a calificar/'")
        return

    os.makedirs(DIR_SALIDA, exist_ok=True)
    print(f"  + Salida: {DIR_SALIDA}")

    todos = []
    for ruta in archivos:
        print(f"\n  Procesando: {os.path.basename(ruta)}")
        alumnos = procesar(ruta, respuestas)
        if not alumnos:
            print("    -> Sin datos válidos")
            continue
        for al in alumnos:
            print(f"    {al['nombre']:<35} GRADO {al['grado']:<2} {al['materia']:<12} "
                  f"{al['correctas']}/{al['total']} ({al['pct']}%)")
        todos.extend(alumnos)

    if todos:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        salida = os.path.join(DIR_SALIDA, f"resultados_{ts}.xlsx")
        generar_excel(todos, salida)
        print(f"\n  + Resultados guardados: {salida}")
        print(f"  + Total estudiantes: {len(todos)}")

    print(f"\n{'='*55}")
    print("  LISTO (sin afectar acumulados ni GitHub)")
    print(f"{'='*55}")

if __name__ == "__main__":
    main()
