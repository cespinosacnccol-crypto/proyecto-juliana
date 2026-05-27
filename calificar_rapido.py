"""
CALIFICADOR RÁPIDO — solo pide un Excel, lo califica y muestra resultados.
"""
import os, re, tkinter as tk
from tkinter import filedialog
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

BASE = os.path.dirname(os.path.abspath(__file__))
RUTA_RESPUESTAS = os.path.join(BASE, "RESPUESTAS CORRECTAS PROYECTO JULIANA.xlsx")
DIR_RESULTADOS = os.path.join(BASE, "Resultados calificación rápida")

FILL_HEADER = PatternFill("solid", fgColor="1F3864")
FONT_HEADER = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
FILL_OK = PatternFill("solid", fgColor="C6EFCE")
FILL_FAIL = PatternFill("solid", fgColor="FFC7CE")
FONT_CELL = Font(size=10, name="Calibri")
ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

SUJETO_MAP = {"lenguaje": "LENGUAJE", "lectura": "LENGUAJE", "matematicas": "MATEMÁTICAS"}
PARED = re.compile(r'[^a-zA-Z0-9áéíóúÁÉÍÓÚñÑüÜ\s]')

def normalizar(v):
    return re.sub(r'\s+', ' ', PARED.sub('', str(v).strip().upper())).strip()

def cargar_respuestas():
    wb = openpyxl.load_workbook(RUTA_RESPUESTAS)
    r = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]: continue
            g = str(row[0]).strip()
            m = str(row[1]).strip().lower()
            for c in SUJETO_MAP:
                if c in m: m = SUJETO_MAP[c]; break
            p = str(row[2]).strip()
            r[(g, m, p)] = str(row[3] or "").strip().upper()
    wb.close()
    return r

def detectar(ws):
    dc = {"NOMBRES ESTUDIANTE": None, "CÓDIGO DANE SEDE": None,
          "NOMBRE SEDE": None, "CÓD. EST.": None, "GRADO": None, "CURSO": None, "PRUEBA": None}
    ini = None
    for c in range(1, ws.max_column + 1):
        v = re.sub(r'\s+', ' ', str(ws.cell(1, c).value or "").strip().upper())
        for k in dc:
            if k in v: dc[k] = c
        if v.startswith("P") and v[1:2].isdigit() and not v.endswith("_EVAL") and ini is None:
            ini = c
    return dc, ini or 1

def procesar(ruta):
    wb = openpyxl.load_workbook(ruta)
    ws = wb.active
    cols, ini = detectar(ws)

    if cols["CÓDIGO DANE SEDE"] is None or cols["NOMBRES ESTUDIANTE"] is None:
        print("  [!] No se detectaron las columnas estándar en el archivo")
        wb.close()
        return []

    respuestas = cargar_respuestas()
    alumnos = []
    for fila in ws.iter_rows(min_row=2, values_only=True):
        nombre = normalizar(fila[cols["NOMBRES ESTUDIANTE"] - 1]) if cols["NOMBRES ESTUDIANTE"] and len(fila) >= cols["NOMBRES ESTUDIANTE"] else ""
        if not nombre: continue
        materia_raw = normalizar(fila[cols["PRUEBA"] - 1]) if cols["PRUEBA"] and len(fila) >= cols["PRUEBA"] else ""
        materia = "LENGUAJE" if "LENGUAJE" in materia_raw or "LECTURA" in materia_raw else "MATEMÁTICAS"
        grado = normalizar(fila[cols["GRADO"] - 1]) if cols["GRADO"] and len(fila) >= cols["GRADO"] else ""

        detalles = []
        for q in range(20):
            col = ini + q - 1
            if col >= len(fila): break
            ra = normalizar(fila[col]) if fila[col] else ""
            clave = (grado, materia, f"P{q + 1:02d}")
            ok = ra == normalizar(respuestas.get(clave, "")) if ra else False
            detalles.append({"resp": ra or "", "ok": ok})
        c = sum(1 for d in detalles if d["ok"])
        t = len(detalles)
        alumnos.append({"nombre": nombre, "grado": grado, "materia": materia,
                       "detalles": detalles, "correctas": c, "total": t,
                       "pct": round(c / t * 100, 1) if t else 0})
    wb.close()
    return alumnos

def guardar(alumnos, ruta):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultados"
    headers = ["NOMBRES ESTUDIANTE", "GRADO", "PRUEBA"] + \
              [f"P{q:02d}" for q in range(1, 21)] + \
              [f"P{q:02d}_EVAL" for q in range(1, 21)] + \
              ["CORRECTAS", "INCORRECTAS", "TOTAL", "PORCENTAJE ACIERTO"]
    for c, h in enumerate(headers, 1):
        celda = ws.cell(1, c, h)
        celda.fill = FILL_HEADER
        celda.font = FONT_HEADER
        celda.alignment = ALIGN
        celda.border = BORDER
    for i, al in enumerate(alumnos, 2):
        ws.cell(i, 1, al["nombre"]).font = FONT_CELL
        ws.cell(i, 2, al["grado"]).font = FONT_CELL
        ws.cell(i, 3, al["materia"]).font = FONT_CELL
        for c in range(1, 4):
            ws.cell(i, c).alignment = ALIGN
            ws.cell(i, c).border = BORDER
        for q, d in enumerate(al["detalles"]):
            cr = 4 + q * 2
            ce = cr + 1
            ws.cell(i, cr, d["resp"]).font = FONT_CELL
            ws.cell(i, cr).alignment = ALIGN
            ws.cell(i, cr).border = BORDER
            celda = ws.cell(i, ce, "CORRECTO" if d["ok"] else "INCORRECTO")
            celda.font = FONT_CELL
            celda.alignment = ALIGN
            celda.border = BORDER
            celda.fill = FILL_OK if d["ok"] else FILL_FAIL
        ws.cell(i, 44, al["correctas"]).font = FONT_CELL
        ws.cell(i, 45, al["total"] - al["correctas"]).font = FONT_CELL
        ws.cell(i, 46, al["total"]).font = FONT_CELL
        ws.cell(i, 47, f"{al['pct']}%").font = FONT_CELL
    wb.save(ruta)
    wb.close()

def main():
    print("=" * 55)
    print("  CALIFICADOR RÁPIDO")
    print("  Selecciona un archivo Excel y lo califica")
    print("=" * 55)

    root = tk.Tk()
    root.withdraw()
    ruta = filedialog.askopenfilename(
        title="Selecciona el archivo Excel de la prueba",
        filetypes=[("Archivos Excel", "*.xlsx"), ("Todos", "*.*")]
    )
    root.destroy()
    if not ruta:
        print("\n  [!] No se seleccionó ningún archivo")
        return

    print(f"\n  Archivo: {os.path.basename(ruta)}")
    alumnos = procesar(ruta)

    if not alumnos:
        print("\n  [!] No se encontraron estudiantes válidos")
        return

    print(f"\n  Estudiantes: {len(alumnos)}")
    for al in alumnos:
        print(f"    {al['nombre']:<35} GRADO {al['grado']:<2} {al['materia']:<12} "
              f"{al['correctas']}/{al['total']} ({al['pct']}%)")

    os.makedirs(DIR_RESULTADOS, exist_ok=True)
    salida = os.path.join(DIR_RESULTADOS,
        os.path.splitext(os.path.basename(ruta))[0] + "_CALIFICADO.xlsx")
    guardar(alumnos, salida)
    print(f"  + Guardado en: {DIR_RESULTADOS}")
    print(f"\n{'='*55}")
    print("  LISTO (sin tocar acumulados ni GitHub)")
    print(f"{'='*55}")

if __name__ == "__main__":
    main()
