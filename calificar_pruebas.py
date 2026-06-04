"""
SISTEMA AUTOMÁTICO DE CALIFICACIÓN DE PRUEBAS — PROYECTO JULIANA
================================================================

1. Coloca los archivos .xlsx de pruebas en:   Pruebas a calificar/
2. Ejecuta: python calificar_pruebas.py
3. El sistema:
   - Lee cada prueba y detecta automáticamente los datos
   - Busca las respuestas correctas según grado y materia
   - Califica cada respuesta (Correcto/Incorrecto con colores)
   - Genera fórmulas Excel (COUNTIF) para resumen
   - Guarda el archivo calificado en:          Pruebas calificadas/
   - Mueve el original a:                       Pruebas a calificar/Ya procesadas/
   - Acumula todo el histórico en:              Pruebas acumuladas/ACUMULADO GENERAL.xlsx
"""

import os, shutil, unicodedata, re, subprocess
from datetime import datetime

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════════════════════
#  RUTAS
# ═══════════════════════════════════════════════════════════════════════════════

BASE            = os.path.dirname(os.path.abspath(__file__))
RUTA_RESPUESTAS = os.path.join(BASE, "RESPUESTAS CORRECTAS PROYECTO JULIANA.xlsx")
DIR_IN          = os.path.join(BASE, "Pruebas a calificar")
DIR_OUT         = os.path.join(BASE, "Pruebas calificadas")
DIR_ACUM        = os.path.join(BASE, "Pruebas acumuladas")
DIR_PROCESADOS  = os.path.join(DIR_IN, "Ya procesadas")
DIR_INFORME     = os.path.join(BASE, "INFORME CLIENTE")
ACUMULADO       = os.path.join(DIR_ACUM, "ACUMULADO GENERAL.xlsx")

# ─── Estilos (igual que los ejemplos) ──────────────────────────────────────────
ST_HDR_FILL   = PatternFill("solid", fgColor="1F3864")
ST_HDR_FONT   = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
ST_HDR_ALIGN  = Alignment(horizontal="center", vertical="center")
ST_CORR_FILL  = PatternFill("solid", fgColor="C6EFCE")
ST_CORR_FONT  = Font(bold=True, color="375623", name="Calibri", size=10)
ST_INCOR_FILL = PatternFill("solid", fgColor="FFC7CE")
ST_INCOR_FONT = Font(bold=True, color="9C0006", name="Calibri", size=10)
ST_CELL_FONT  = Font(name="Calibri", size=10)
ST_CELL_ALIGN = Alignment(horizontal="center", vertical="center")
ST_THIN       = Side(style="thin", color="BFBFBF")
ST_BORDER     = Border(left=ST_THIN, right=ST_THIN, top=ST_THIN, bottom=ST_THIN)
ST_SUM_FONT   = Font(bold=True, name="Calibri", size=10)

# Anchocolumnas por posición (1-indexed) para metadatos y por nombre para el resto
ANCHOS_META = {1: 10, 2: 18, 3: 30, 4: 34, 5: 14, 6: 8, 7: 8, 8: 14}
ANCHOS_SUM = {
    "CORRECTAS": 12, "INCORRECTAS": 12, "TOTAL_EVAL": 12, "PORCENTAJE ACIERTO": 18,
}

MATERIA_NOMBRES = {
    "matematicas": "MATEMÁTICAS",
    "lenguaje": "LENGUAJE",
}


# ═══════════════════════════════════════════════════════════════════════════════
#  UTILIDADES
# ═══════════════════════════════════════════════════════════════════════════════

def norm(t):
    """Minúsculas, sin tildes, sin espacios extra."""
    if t is None: return ""
    t = unicodedata.normalize("NFD", str(t).strip().lower())
    return "".join(c for c in t if unicodedata.category(c) != "Mn")


def crear_carpetas():
    for p in [DIR_IN, DIR_OUT, DIR_ACUM, DIR_PROCESADOS, DIR_INFORME]:
        os.makedirs(p, exist_ok=True)


def sanitizar_sheet(texto):
    """Nombre de pestaña válido para Excel (máx 31 caracteres), conserva tildes."""
    s = texto.strip().upper().replace(" ", "_")
    s = re.sub(r"[\[\]:\*\?/\\]", "", s)
    return s[:31]


def estilo_header(ws, total_cols):
    for c in range(1, total_cols + 1):
        cell = ws.cell(1, c)
        cell.fill = ST_HDR_FILL
        cell.font = ST_HDR_FONT
        cell.alignment = ST_HDR_ALIGN
        cell.border = ST_BORDER
    ws.row_dimensions[1].height = 28


def ancho_columnas(ws, headers):
    n_meta = len(COLUMNAS_ESTANDAR)
    for i, h in enumerate(headers, 1):
        if i <= n_meta:
            w = ANCHOS_META.get(i, 13)
        else:
            w = ANCHOS_SUM.get(h, 13)
        ws.column_dimensions[get_column_letter(i)].width = w


def colorear_eval(cell, valor):
    """Aplica color según CORRECTO/INCORRECTO (todo mayúsculas)."""
    cell.value = valor.upper()
    cell.border = ST_BORDER
    cell.alignment = ST_CELL_ALIGN
    if valor.upper() == "CORRECTO":
        cell.fill = ST_CORR_FILL
        cell.font = ST_CORR_FONT
    else:
        cell.fill = ST_INCOR_FILL
        cell.font = ST_INCOR_FONT


# ═══════════════════════════════════════════════════════════════════════════════
#  CARGAR MAPEO TRATADO/CONTROL DESDE SEGUIMIENTO
# ═══════════════════════════════════════════════════════════════════════════════

RUTA_SEGUIMIENTO = os.path.join(BASE, "Seguimiento.xlsx")

def cargar_tipo_mapping():
    """Retorna dict: cod_dane -> 'TRATADO'/'CONTROL'/'DESCONOCIDO'"""
    mapping = {}
    if not os.path.exists(RUTA_SEGUIMIENTO):
        print("  [!] No se encuentra Seguimiento.xlsx — todos serán DESCONOCIDO")
        return mapping
    wb = openpyxl.load_workbook(RUTA_SEGUIMIENTO)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        cod = str(row[0]).strip() if row[0] else ""
        tipo = str(row[3]).strip().upper() if len(row) > 3 and row[3] else ""
        if cod and tipo:
            mapping[cod] = tipo
    wb.close()
    print(f"  + Mapeo Tratado/Control cargado: {len(mapping)} sedes")
    return mapping

TIPO_MAPPING = cargar_tipo_mapping()

# ═══════════════════════════════════════════════════════════════════════════════
#  CARGAR RESPUESTAS CORRECTAS
# ═══════════════════════════════════════════════════════════════════════════════

def cargar_respuestas():
    wb = openpyxl.load_workbook(RUTA_RESPUESTAS)
    ws = wb.active
    resp = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        g = int(row[0])
        m = norm(row[1])
        n = int(row[2])
        r = str(row[3]).strip().upper() if row[3] else ""
        resp.setdefault((g, m), {})[n] = r
    print(f"\n  + Respuestas cargadas: {len(resp)} combinaciones")
    for (g, m), v in sorted(resp.items()):
        print(f"    - Grado {g} / {m.upper()}: {len(v)} preguntas")
    return resp


# ═══════════════════════════════════════════════════════════════════════════════
#  DETECTAR COLUMNAS EN ARCHIVO DE ENTRADA (versión mejorada)
# ═══════════════════════════════════════════════════════════════════════════════

COLUMNAS_ESTANDAR = [
    ("TIPO",              "tipo"),
    ("CÓDIGO DANE SEDE",  "cod_dane"),
    ("NOMBRE SEDE",       "nombre_sede"),
    ("NOMBRES ESTUDIANTE", "estudiante"),
    ("CÓD. EST.",         "id"),
    ("GRADO",             "grado"),
    ("CURSO",             "curso"),
    ("PRUEBA",            "materia"),
]


def _analizar_muestras(ws, col_idx, max_muestras=5):
    """Analiza valores de muestra de una columna para inferir su tipo.
    Retorna una de: 'dane', 'sede', 'id', 'nombre', None."""
    muestras = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if i >= max_muestras:
            break
        v = row[col_idx] if col_idx < len(row) else None
        if v is not None and str(v).strip():
            muestras.append(str(v).strip())

    if not muestras:
        return None

    # 1) Código DANE: 9-13 dígitos, solo números
    if all(v.isdigit() and 9 <= len(v) <= 13 for v in muestras):
        return "dane"

    # 2) Nombre sede: contiene palabras clave de institución educativa
    pal_sede = ["colegio", "ie", "institución", "institucion", "escuela",
                "sede educativa", "centro educativo", "liceo", "jardin", "jardín",
                "normal ", "técnico", "tecnico"]
    if any(any(p in v.lower() for p in pal_sede) for v in muestras):
        return "sede"

    # 3) ID: mezcla letras+números, o 5-11 dígitos puros
    if any(any(c.isdigit() for c in v) and any(c.isalpha() for c in v) for v in muestras):
        return "id"
    if all(v.isdigit() and 5 <= len(v) <= 11 for v in muestras):
        return "id"

    # 4) Nombre estudiante: ≥2 palabras, sin dígitos
    if all(len(v.split()) >= 2 and not any(c.isdigit() for c in v) for v in muestras):
        return "nombre"

    return None


def detectar_original(headers_raw, ws=None):
    """Detecta columnas equivalentes en los encabezados originales
    (y opcionalmente analiza datos de muestra como fallback).
    Retorna dict: key_estandar -> índice 0-based en el archivo original."""
    hn = [norm(h) for h in headers_raw]
    idx = {}
    usado = set()

    def marcar(clave, i):
        idx[clave] = i
        usado.add(i)

    def buscar_exacto(candidatos):
        for cand in candidatos:
            nc = norm(cand)
            for i, h in enumerate(hn):
                if h == nc and i not in usado:
                    return i
        return None

    def buscar_parcial(candidatos, excluir=None):
        for cand in candidatos:
            nc = norm(cand)
            for i, h in enumerate(hn):
                if i in usado:
                    continue
                if excluir and i in excluir:
                    continue
                if nc in h:
                    return i
        return None

    # ── 1. Coincidencias exactas ──────────────────────────────────────────
    # DANE
    i = buscar_exacto(["codigo dane sede", "codigo dane", "codigo sede",
                        "cod dane", "codigo institucion", "codigo sede",
                        "dane", "id sede", "id_sede"])
    if i is not None: marcar("cod_dane", i)

    # NOMBRE SEDE
    i = buscar_exacto(["nombre sede", "nombre institucion", "nombre colegio",
                        "sede educativa", "institucion educativa"])
    if i is not None: marcar("nombre_sede", i)
    i = buscar_exacto(["colegio", "institucion", "institución", "escuela", "sede"])
    if i is not None: marcar("nombre_sede", i)

    # ESTUDIANTE
    i = buscar_exacto(["nombres estudiante", "nombre del estudiante",
                        "nombre estudiante", "nombre completo", "estudiante",
                        "nombres", "alumno"])
    if i is not None: marcar("estudiante", i)

    # ID
    i = buscar_exacto(["id", "id estudiante", "cod est", "cód est",
                        "cod. est", "cód. est",
                        "cod.est", "cod.est.", "codest",
                        "codigo estudiante", "código estudiante",
                        "codigo est", "código est", "codigo",
                        "documento", "identificacion",
                        "identificación", "cedula", "cédula", "nuip",
                        "tarjeta identidad", "numero documento"])
    if i is not None: marcar("id", i)

    # GRADO
    i = buscar_exacto(["grado"])
    if i is not None: marcar("grado", i)

    # CURSO
    i = buscar_exacto(["curso", "grupo"])
    if i is not None: marcar("curso", i)

    # MATERIA
    i = buscar_exacto(["prueba", "materia", "asignatura", "lectura"])
    if i is not None: marcar("materia", i)

    # ── 2. Coincidencias parciales ────────────────────────────────────────
    # DANE (evitar que "dane" genérico enganche "nombre sede" u otras)
    if "cod_dane" not in idx:
        i = buscar_parcial(["codigo dane", "codigo sede", "cod dane", "dane"],
                           excluir={idx.get("nombre_sede")})
        if i is not None: marcar("cod_dane", i)

    # NOMBRE SEDE
    if "nombre_sede" not in idx:
        i = buscar_parcial(["nombre sede", "nombre institucion", "nombre colegio",
                            "sede educativa", "colegio", "institucion", "institución",
                            "escuela", "sede", "nombre institucion", "nombre colegio",
                            "ie"])
        if i is not None:
            # Si apunta a la misma columna que cod_dane, descartar
            if idx.get("cod_dane") is not None and i == idx["cod_dane"]:
                pass  # No asignar
            else:
                marcar("nombre_sede", i)

    # ESTUDIANTE
    if "estudiante" not in idx:
        i = buscar_parcial(["nombres estudiante", "estudiante", "alumno",
                            "nombre completo", "nombre"])
        if i is not None: marcar("estudiante", i)

    # ID
    if "id" not in idx:
        i = buscar_parcial(["id", "codigo", "código", "documento", "identificacion",
                            "identificación", "cedula", "cédula", "nuip",
                            "tarjeta identidad"])
        if i is not None:
            # Si "codigo" cayó en la columna dane, descartar
            if idx.get("cod_dane") is not None and i == idx["cod_dane"]:
                pass
            else:
                marcar("id", i)

    # GRADO
    if "grado" not in idx:
        i = buscar_parcial(["grado"])
        if i is not None: marcar("grado", i)

    # CURSO
    if "curso" not in idx:
        i = buscar_parcial(["curso", "grupo"])
        if i is not None: marcar("curso", i)

    # MATERIA
    if "materia" not in idx:
        i = buscar_parcial(["prueba", "materia", "asignatura", "lectura"])
        if i is not None: marcar("materia", i)

    # ── 3. Análisis de datos (fallback) ───────────────────────────────────
    if ws is not None:
        # Columnas aún sin detectar
        pendientes = [k for k in ["cod_dane", "nombre_sede", "estudiante", "id"]
                      if k not in idx]
        if pendientes:
            for col_i in range(len(headers_raw)):
                if col_i in usado:
                    continue
                tipo = _analizar_muestras(ws, col_i)
                if tipo == "dane" and "cod_dane" not in idx:
                    marcar("cod_dane", col_i)
                elif tipo == "sede" and "nombre_sede" not in idx:
                    marcar("nombre_sede", col_i)
                elif tipo == "id" and "id" not in idx:
                    marcar("id", col_i)
                elif tipo == "nombre" and "estudiante" not in idx:
                    marcar("estudiante", col_i)

    return idx


def estandarizar_columnas(ws):
    """Reorganiza las columnas del worksheet al orden estándar fijo
    y renombra los encabezados. Devuelve un nuevo workbook.
    Solo mueve las 7 columnas principales; el resto se conserva intacto
    en el orden original."""
    headers_raw = [str(c.value) if c.value is not None else "" for c in ws[1]]
    col = detectar_original(headers_raw, ws=ws)

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active

    orden_claves = [k for _, k in COLUMNAS_ESTANDAR]
    total_meta = len(COLUMNAS_ESTANDAR)

    # Índices de las columnas estándar detectadas
    idx_std = {col.get(k) for k in orden_claves if col.get(k) is not None}

    # ── Encabezados ──
    for i, (nombre, _) in enumerate(COLUMNAS_ESTANDAR, 1):
        ws_out.cell(1, i, nombre)

    out_col = total_meta + 1
    for orig_i, h in enumerate(headers_raw):
        if orig_i not in idx_std:
            ws_out.cell(1, out_col, h)
            out_col += 1

    # ── Datos ──
    for fila_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if all(c is None or str(c).strip() == "" for c in row):
            continue

        # Columnas estándar (1..total_meta)
        for c_idx, key in enumerate(orden_claves, 1):
            src = col.get(key)
            if src is not None and src < len(row):
                ws_out.cell(fila_idx, c_idx, row[src])

        # Columnas no estándar, en orden original
        out_col = total_meta + 1
        for orig_i in range(len(row)):
            if orig_i not in idx_std:
                ws_out.cell(fila_idx, out_col, row[orig_i])
                out_col += 1

    return wb_out


# ═══════════════════════════════════════════════════════════════════════════════
#  PROCESAR UN ARCHIVO
# ═══════════════════════════════════════════════════════════════════════════════

def procesar(ruta, resp_correctas):
    nom = os.path.basename(ruta)
    print(f"\n{'='*55}")
    print(f"  PROCESANDO: {nom}")
    print(f"{'='*55}")

    try:
        wb_in = openpyxl.load_workbook(ruta)
        ws_in = wb_in.active
    except Exception as e:
        print(f"  [!] ERROR: No se pudo abrir ({e})")
        return

    # ── Estandarizar columnas al orden fijo ────────────────────────────────────
    wb_std = estandarizar_columnas(ws_in)
    ws = wb_std.active

    headers = [str(c.value) if c.value is not None else "" for c in ws[1]]
    n_meta = len(COLUMNAS_ESTANDAR)

    # Columnas fijas (0-based después de estandarizar):
    # [0]=TIPO, [1]=Código Dane Sede, [2]=Nombre Sede, [3]=Nombres Estudiante,
    # [4]=Cód. Est., [5]=Grado, [6]=Curso, [7]=Prueba, [8..]=respuestas
    e_tipo       = 0
    e_cod_dane   = 1
    e_nom_sede   = 2
    e_estudiante = 3
    e_id         = 4
    e_grado      = 5
    e_grupo      = 6
    e_materia    = 7
    e_resp       = list(range(n_meta, len(headers)))

    print(f"  + Columnas estandarizadas: {len(e_resp)} respuestas")

    # ── Leer filas ────────────────────────────────────────────────────────────
    alumnos = []
    for fila_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if all(c is None or str(c).strip() == "" for c in row):
            continue

        cod_dane   = str(row[e_cod_dane]).strip() if len(row) > e_cod_dane and row[e_cod_dane] else ""
        tipo       = TIPO_MAPPING.get(cod_dane, "")
        nombre_est = re.sub(r'\s+', ' ', str(row[e_estudiante]).strip().upper()) if len(row) > e_estudiante and row[e_estudiante] else "SIN NOMBRE"
        sede       = re.sub(r'\s+', ' ', str(row[e_nom_sede]).strip().upper()) if len(row) > e_nom_sede and row[e_nom_sede] else os.path.splitext(nom)[0]
        cod_est    = re.sub(r'\s+', ' ', str(row[e_id]).strip().upper()) if len(row) > e_id and row[e_id] else ""
        grupo      = re.sub(r'\s+', ' ', str(row[e_grupo]).strip().upper()) if len(row) > e_grupo and row[e_grupo] else ""

        grado = None
        if e_grado is not None and len(row) > e_grado:
            try: grado = int(float(str(row[e_grado]).strip()))
            except: pass

        materia_raw = ""
        if e_materia is not None and len(row) > e_materia and row[e_materia]:
            materia_raw = norm(str(row[e_materia]))
            if materia_raw == "lectura":
                materia_raw = "lenguaje"

        # Fallback: extraer grado/materia del nombre del archivo
        if grado is None or not materia_raw:
            nb = os.path.splitext(nom)[0].replace(" ", "_").replace("-", "_")
            for p in nb.split("_"):
                pu = norm(p).upper()
                if grado is None and pu in ("3","5","7","9","11"):
                    grado = int(pu)
                if not materia_raw:
                    if pu in ("MATEMATICAS","MATE","MAT"):
                        materia_raw = "matematicas"
                    elif pu in ("LENGUAJE","LENGUA","LEN"):
                        materia_raw = "lenguaje"

        if grado is None:
            print(f"  [!] Fila {fila_idx}: No se detectó GRADO → salta")
            continue
        if not materia_raw:
            print(f"  [!] Fila {fila_idx}: No se detectó MATERIA → salta")
            continue

        key = (grado, materia_raw)
        if key not in resp_correctas:
            print(f"  [!] Fila {fila_idx}: Sin respuestas para Grado {grado}/{materia_raw.upper()} → salta")
            continue

        correctas_dict = resp_correctas[key]
        total_preg = len(correctas_dict)

        # Extraer respuestas
        resp_est = []
        for idx_col in e_resp:
            val = row[idx_col] if idx_col < len(row) else None
            resp_est.append(str(val).strip().upper() if val is not None else "")

        # Calificar
        detalles = []
        num_ok = 0
        for i in range(total_preg):
            num_p = i + 1
            r_est = resp_est[i] if i < len(resp_est) else ""
            r_cor = correctas_dict.get(num_p, "")
            ok = (r_est == r_cor and r_est != "")
            if ok: num_ok += 1
            detalles.append({"resp": r_est, "correcta": r_cor, "ok": ok})

        num_bad = total_preg - num_ok
        pct = round(num_ok / total_preg, 4) if total_preg else 0

        alumnos.append({
            "tipo": tipo, "cod_dane": cod_dane, "sede": sede, "estudiante": nombre_est, "cod": cod_est,
            "grado": grado, "grupo": grupo, "materia": materia_raw,
            "detalles": detalles, "total": total_preg,
            "ok": num_ok, "bad": num_bad, "pct": pct,
        })

        print(f"  + {nombre_est[:35].upper():<35} GRADO {grado} {materia_raw.upper():<12} "
              f"{num_ok}/{total_preg} ({round(pct*100,1)}%)")

    if not alumnos:
        print("  [!] No se encontraron estudiantes válidos")
        return

    total_preg = alumnos[0]["total"]
    nom_sin_ext = os.path.splitext(nom)[0]

    # ═══════════════════════════════════════════════════════════════════════════
    #  GENERAR ARCHIVO CALIFICADO  (formato = ejemplo lenguaje)
    # ═══════════════════════════════════════════════════════════════════════════

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Evaluaciones"

    # Construir encabezados
    cols_meta = ["TIPO", "CÓDIGO DANE SEDE", "NOMBRE SEDE", "NOMBRES ESTUDIANTE",
                  "CÓD. EST.", "GRADO", "CURSO", "PRUEBA"]
    cols_preg = []
    for q in range(1, total_preg + 1):
        cols_preg.append(f"P{q:02d}")
        cols_preg.append(f"P{q:02d}_EVAL")
    cols_sum = ["CORRECTAS", "INCORRECTAS", "TOTAL_EVAL", "PORCENTAJE ACIERTO"]
    todos_cols = cols_meta + cols_preg + cols_sum

    for c, h in enumerate(todos_cols, 1):
        ws.cell(1, c, h)
    estilo_header(ws, len(todos_cols))
    ancho_columnas(ws, todos_cols)

    # Rango para COUNTIF
    col_ini_eval = len(cols_meta) + 2   # primera columna _eval
    col_fin_eval = len(cols_meta) + total_preg * 2  # última columna _eval
    letra_ini = get_column_letter(col_ini_eval)
    letra_fin = get_column_letter(col_fin_eval)

    col_idx_corr  = todos_cols.index("CORRECTAS") + 1
    col_idx_inc   = todos_cols.index("INCORRECTAS") + 1
    col_idx_total = todos_cols.index("TOTAL_EVAL") + 1
    col_idx_pct   = todos_cols.index("PORCENTAJE ACIERTO") + 1

    for i, al in enumerate(alumnos):
        fila = 2 + i
        ws.cell(fila, 1, al["tipo"]).font = ST_CELL_FONT
        ws.cell(fila, 2, al["cod_dane"]).font = ST_CELL_FONT
        ws.cell(fila, 3, al["sede"].upper()).font = ST_CELL_FONT
        ws.cell(fila, 4, al["estudiante"].upper()).font = ST_CELL_FONT
        ws.cell(fila, 5, al["cod"].upper()).font = ST_CELL_FONT
        ws.cell(fila, 6, al["grado"]).font = ST_CELL_FONT
        gpo = int(al["grupo"]) if al["grupo"].isdigit() else al["grupo"].upper()
        ws.cell(fila, 7, gpo).font = ST_CELL_FONT
        ws.cell(fila, 8, MATERIA_NOMBRES.get(al["materia"], al["materia"].upper())).font = ST_CELL_FONT

        for c in range(1, 9):
            cell = ws.cell(fila, c)
            cell.alignment = ST_CELL_ALIGN
            cell.border = ST_BORDER

        for q, det in enumerate(al["detalles"]):
            col_resp = len(cols_meta) + 1 + q * 2
            col_eval = col_resp + 1

            # Respuesta
            c_resp = ws.cell(fila, col_resp, det["resp"])
            c_resp.font = ST_CELL_FONT
            c_resp.alignment = ST_CELL_ALIGN
            c_resp.border = ST_BORDER

            # Evaluación
            c_eval = ws.cell(fila, col_eval)
            colorear_eval(c_eval, "CORRECTO" if det["ok"] else "INCORRECTO")

        # Fórmulas de resumen
        c_corr = ws.cell(fila, col_idx_corr)
        c_corr.value = f'=COUNTIF({letra_ini}{fila}:{letra_fin}{fila},"CORRECTO")'
        c_corr.font = ST_SUM_FONT; c_corr.alignment = ST_CELL_ALIGN; c_corr.border = ST_BORDER

        c_inc = ws.cell(fila, col_idx_inc)
        c_inc.value = f'=COUNTIF({letra_ini}{fila}:{letra_fin}{fila},"INCORRECTO")'
        c_inc.font = ST_SUM_FONT; c_inc.alignment = ST_CELL_ALIGN; c_inc.border = ST_BORDER

        c_tot = ws.cell(fila, col_idx_total, total_preg)
        c_tot.font = ST_SUM_FONT; c_tot.alignment = ST_CELL_ALIGN; c_tot.border = ST_BORDER

        c_pct = ws.cell(fila, col_idx_pct)
        c_pct.value = f'={get_column_letter(col_idx_corr)}{fila}/{get_column_letter(col_idx_total)}{fila}'
        c_pct.font = ST_SUM_FONT; c_pct.alignment = ST_CELL_ALIGN; c_pct.border = ST_BORDER
        c_pct.number_format = "0.00%"

    ws.freeze_panes = "A2"

    # Guardar calificado
    ruta_sal = os.path.join(DIR_OUT, f"{nom_sin_ext}_CALIFICADO.xlsx")
    try:
        wb.save(ruta_sal)
        print(f"\n  + Guardado: {ruta_sal}")
    except Exception as e:
        print(f"  [!] ERROR al guardar: {e}")
        return

    # ── Mover original a Ya procesadas ─────────────────────────────────────────
    try:
        # Renombrar con fecha/hora para mantener histórico
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        nom_historico = f"{nom_sin_ext}_{ts}.xlsx"
        shutil.move(ruta, os.path.join(DIR_PROCESADOS, nom_historico))
        print(f"  + Original movido a: Ya procesadas/ ({nom_historico})")
    except Exception as e:
        print(f"  [!] ERROR al mover original: {e}")

    # ── Actualizar acumulado ───────────────────────────────────────────────────
    actualizar_acumulado(alumnos)


# ═══════════════════════════════════════════════════════════════════════════════
#  ACTUALIZAR ACUMULADO GENERAL
# ═══════════════════════════════════════════════════════════════════════════════

def _sort_key_grado(name):
    """Ordena sheets por grado ascendente (3, 5, 7, 9, 11)."""
    return int(name.split("_")[0])


def _reordenar_sheets(wb):
    """Reordena las hojas del workbook por grado ascendente."""
    target = sorted(wb.sheetnames, key=_sort_key_grado)
    for i, name in enumerate(target):
        cur = wb.sheetnames.index(name)
        if cur != i:
            wb.move_sheet(name, offset=i - cur)


def _append_alumnos_to_wb(wb, alumnos, total_preg):
    """Agrega alumnos a un workbook acumulado. Retorna el workbook."""
    grupos = {}
    for al in alumnos:
        mat_display = MATERIA_NOMBRES.get(al["materia"], al["materia"].upper())
        key = f"{al['grado']}_{sanitizar_sheet(mat_display)}"
        grupos.setdefault(key, []).append(al)

    cols_meta = ["TIPO", "CÓDIGO DANE SEDE", "NOMBRE SEDE", "NOMBRES ESTUDIANTE",
                  "CÓD. EST.", "GRADO", "CURSO", "PRUEBA"]
    cols_preg = []
    for q in range(1, total_preg + 1):
        cols_preg.append(f"P{q:02d}_EVAL")
    cols_sum = ["CORRECTAS", "INCORRECTAS", "TOTAL_EVAL", "PORCENTAJE ACIERTO"]
    todos_cols = cols_meta + cols_preg + cols_sum

    for sheet_name, grupo_alumnos in sorted(
        grupos.items(), key=lambda x: _sort_key_grado(x[0])
    ):
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)
            for c, h in enumerate(todos_cols, 1):
                ws.cell(1, c, h)
            estilo_header(ws, len(todos_cols))
            ancho_columnas(ws, todos_cols)
            ws.freeze_panes = "A2"

        col_ini_eval = len(cols_meta) + 1
        col_fin_eval = len(cols_meta) + total_preg
        letra_ini = get_column_letter(col_ini_eval)
        letra_fin = get_column_letter(col_fin_eval)

        prox_fila = (ws.max_row or 0) + 1

        for al in grupo_alumnos:
            fila = prox_fila
            ws.cell(fila, 1, al["tipo"]).font = ST_CELL_FONT
            ws.cell(fila, 2, al["cod_dane"]).font = ST_CELL_FONT
            ws.cell(fila, 3, al["sede"].upper()).font = ST_CELL_FONT
            ws.cell(fila, 4, al["estudiante"].upper()).font = ST_CELL_FONT
            ws.cell(fila, 5, al["cod"].upper()).font = ST_CELL_FONT
            ws.cell(fila, 6, al["grado"]).font = ST_CELL_FONT
            gpo = int(al["grupo"]) if al["grupo"].isdigit() else al["grupo"].upper()
            ws.cell(fila, 7, gpo).font = ST_CELL_FONT
            ws.cell(fila, 8, MATERIA_NOMBRES.get(al["materia"], al["materia"].upper())).font = ST_CELL_FONT

            for c in range(1, 9):
                ws.cell(fila, c).alignment = ST_CELL_ALIGN
                ws.cell(fila, c).border = ST_BORDER

            for q, det in enumerate(al["detalles"]):
                col_eval = len(cols_meta) + 1 + q
                colorear_eval(ws.cell(fila, col_eval), "CORRECTO" if det["ok"] else "INCORRECTO")

            col_corr = len(todos_cols) - 3
            col_inc  = col_corr + 1
            col_tot  = col_inc + 1
            col_pct  = col_tot + 1

            c_c = ws.cell(fila, col_corr)
            c_c.value = f'=COUNTIF({letra_ini}{fila}:{letra_fin}{fila},"CORRECTO")'
            c_c.font = ST_SUM_FONT; c_c.alignment = ST_CELL_ALIGN; c_c.border = ST_BORDER

            c_i = ws.cell(fila, col_inc)
            c_i.value = f'=COUNTIF({letra_ini}{fila}:{letra_fin}{fila},"INCORRECTO")'
            c_i.font = ST_SUM_FONT; c_i.alignment = ST_CELL_ALIGN; c_i.border = ST_BORDER

            c_t = ws.cell(fila, col_tot, total_preg)
            c_t.font = ST_SUM_FONT; c_t.alignment = ST_CELL_ALIGN; c_t.border = ST_BORDER

            c_p = ws.cell(fila, col_pct)
            c_p.value = f'={get_column_letter(col_corr)}{fila}/{get_column_letter(col_tot)}{fila}'
            c_p.font = ST_SUM_FONT; c_p.alignment = ST_CELL_ALIGN; c_p.border = ST_BORDER
            c_p.number_format = "0.00%"

            prox_fila += 1

        print(f"  + Acumulado [{sheet_name}] +{len(grupo_alumnos)} estudiantes "
              f"(fila {prox_fila - len(grupo_alumnos)} en adelante)")
    _reordenar_sheets(wb)
    return wb


def actualizar_acumulado(alumnos):
    """Agrega los alumnos al archivo acumulado en disco."""
    if not alumnos:
        return
    total_preg = alumnos[0]["total"]
    if os.path.exists(ACUMULADO):
        wb = openpyxl.load_workbook(ACUMULADO)
    else:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
    wb = _append_alumnos_to_wb(wb, alumnos, total_preg)
    _reordenar_sheets(wb)
    try:
        wb.save(ACUMULADO)
        print(f"  + Acumulado guardado: {ACUMULADO}")
    except Exception as e:
        print(f"  [!] ERROR al guardar acumulado: {e}")


def actualizar_acumulado_wb(alumnos, wb):
    """Agrega alumnos a un workbook acumulado en memoria (sin guardar en disco)."""
    if not alumnos:
        return wb
    total_preg = alumnos[0]["total"]
    wb = _append_alumnos_to_wb(wb, alumnos, total_preg)
    _reordenar_sheets(wb)
    return wb


# ═══════════════════════════════════════════════════════════════════════════════
#  GENERAR INFORME CLIENTE (copia del acumulado sin respuestas visibles)
# ═══════════════════════════════════════════════════════════════════════════════

def generar_informe_cliente():
    """Copia ACUMULADO GENERAL.xlsx a INFORME CLIENTE y elimina
    las columnas de respuesta (P01, P02, ...) dejando solo _eval."""
    if not os.path.exists(ACUMULADO):
        return

    destino = os.path.join(DIR_INFORME, "ACUMULADO GENERAL.xlsx")
    shutil.copy2(ACUMULADO, destino)

    wb = openpyxl.load_workbook(destino)
    _reordenar_sheets(wb)
    for ws in wb.worksheets:
        headers = [str(ws.cell(1, c).value or "") for c in range(1, ws.max_column + 1)]
        p_cols = sorted([i+1 for i, h in enumerate(headers) if re.match(r'^P\d{2}$', h)], reverse=True)
        for ci in p_cols:
            ws.delete_cols(ci)
        headers = [str(ws.cell(1, c).value or "") for c in range(1, ws.max_column + 1)]
        eval_idx = [i for i, h in enumerate(headers) if re.match(r'^P\d{2}_EVAL$', h)]
        total = len(eval_idx)
        col_corr = next((i+1 for i, h in enumerate(headers) if h == "CORRECTAS"), None)
        col_inc = next((i+1 for i, h in enumerate(headers) if h == "INCORRECTAS"), None)
        col_pct = next((i+1 for i, h in enumerate(headers) if h == "PORCENTAJE ACIERTO"), None)
        for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if row[0] is None:
                continue
            correctas = sum(1 for i in eval_idx if i < len(row) and str(row[i]).strip().upper() == "CORRECTO")
            if col_corr:
                ws.cell(r_idx, col_corr).value = correctas
            if col_inc:
                ws.cell(r_idx, col_inc).value = total - correctas
            if col_pct:
                ws.cell(r_idx, col_pct).value = round(correctas / total, 4) if total else 0

    wb.save(destino)
    print(f"  + Informe cliente: {destino}")


# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    print("=" * 55)
    print("  SISTEMA DE CALIFICACION AUTOMATICA")
    print("  PROYECTO JULIANA")
    print("=" * 55)

    print("\n[1] Verificando carpetas...")
    crear_carpetas()
    print("  + OK")

    print("\n[2] Cargando respuestas correctas...")
    if not os.path.exists(RUTA_RESPUESTAS):
        print(f"  [!] ERROR: No se encuentra:\n      {RUTA_RESPUESTAS}")
        return
    resp = cargar_respuestas()
    if not resp:
        print("  [!] No hay respuestas. Saliendo.")
        return

    print("\n[3] Buscando pruebas en 'Pruebas a calificar/'...")
    archivos = sorted(
        os.path.join(DIR_IN, f)
        for f in os.listdir(DIR_IN)
        if f.lower().endswith(".xlsx") and os.path.isfile(os.path.join(DIR_IN, f))
    )
    if not archivos:
        print("  [!] No hay archivos .xlsx en 'Pruebas a calificar/'")
        print("      Coloca los archivos y ejecuta de nuevo.")
        return

    print(f"  + {len(archivos)} archivo(s) encontrado(s):")
    for a in archivos:
        print(f"    - {os.path.basename(a)}")

    print(f"\n[4] Calificando...")
    ok = err = 0
    for a in archivos:
        try:
            procesar(a, resp)
            ok += 1
        except Exception as e:
            print(f"  [!] ERROR procesando {os.path.basename(a)}: {e}")
            err += 1

    # ── Generar informe cliente ────────────────────────────────────────────────
    print("\n[5] Generando informe cliente...")
    generar_informe_cliente()

    print(f"\n{'='*55}")
    print(f"  PROCESO COMPLETADO")
    print(f"{'='*55}")
    print(f"  Procesados correctamente: {ok}")
    print(f"  Con error:                {err}")
    print(f"  Total archivos:           {len(archivos)}")

    print(f"\n  Calificados    : {DIR_OUT}")
    print(f"  Acumulado      : {ACUMULADO}")
    print(f"  Informe cliente: {DIR_INFORME}")
    print(f"  Procesados     : {DIR_PROCESADOS}")

    # ── Subir a GitHub ──────────────────────────────────────────
    print("\n[6] Subiendo a GitHub...")
    try:
        subprocess.run(["git", "add", "-A"], cwd=BASE, capture_output=True, text=True)
        r = subprocess.run(["git", "diff", "--cached", "--quiet"], cwd=BASE, capture_output=True)
        if r.returncode != 0:
            msg = f"actualiza acumulados {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            subprocess.run(["git", "commit", "-m", msg], cwd=BASE, capture_output=True, text=True)
            p = subprocess.run(["git", "push"], cwd=BASE, capture_output=True, text=True)
            if p.returncode == 0:
                print("  + Subido a GitHub — Streamlit Cloud se actualizará automáticamente")
            else:
                print(f"  [!] Error al hacer push: {p.stderr.strip()}")
        else:
            print("  + Sin cambios nuevos para subir")
    except Exception as e:
        print(f"  [!] Error en git: {e}")


if __name__ == "__main__":
    main()
